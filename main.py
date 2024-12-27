import os
from flask import Flask, request, redirect, url_for, render_template, send_file, flash
import openpyxl
from collections import defaultdict

# Tạo ứng dụng Flask
app = Flask(__name__)
app.secret_key = "your_secret_key"

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["OUTPUT_FOLDER"] = OUTPUT_FOLDER


# Hàm xử lý file Excel
def convert_survey_data(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    sheet = wb.active

    # Initialize data structures
    respondents = {}
    questions = []
    data = defaultdict(dict)
    emails = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[2]  # Column C: Tên người khảo sát
        email = row[3]  # Column D: Email
        question = row[4]  # Column E: Câu hỏi
        answer = row[5]  # Column F: Câu trả lời

        if question not in questions:
            questions.append(question)

        respondents[name] = name
        emails[name] = email
        data[name][question] = answer

    # Create a new workbook and select the active sheet
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    new_sheet.title = "Converted Survey Data"

    # Write headers
    headers = ["Tên người khảo sát", "Email"] + questions
    new_sheet.append(headers)

    # Write data
    for name in respondents:
        row_data = [name, emails.get(name, "")]
        for question in questions:
            row_data.append(data[name].get(question, ""))
        new_sheet.append(row_data)

    # Save the new workbook
    new_wb.save(output_path)


# Trang chủ - Upload file
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # Kiểm tra file upload
        if "file" not in request.files:
            flash("No file part")
            return redirect(request.url)
        file = request.files["file"]
        if file.filename == "":
            flash("No selected file")
            return redirect(request.url)
        if file:
            input_path = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
            output_path = os.path.join(app.config["OUTPUT_FOLDER"], f"converted_{file.filename}")
            file.save(input_path)

            # Gọi hàm xử lý file
            convert_survey_data(input_path, output_path)
            flash("File processed successfully!")
            return redirect(url_for("download_file", filename=f"converted_{file.filename}"))

    return render_template("index.html")


# Trang download file
@app.route("/download/<filename>")
def download_file(filename):
    return send_file(os.path.join(app.config["OUTPUT_FOLDER"], filename), as_attachment=True)


# Chạy ứng dụng Flask
if __name__ == "__main__":
    app.run(debug=True)
